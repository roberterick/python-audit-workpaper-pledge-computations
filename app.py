import decimal
import itertools
import library
import pledgemod
import openpyxl as opx
import os
ZERO=decimal.Decimal()
ALL_DATA=True
OUTPUT='result.xlsx'

class App:
    def __init__(self):
        print('initializing parent')
        self.header=['pledge_number','account_number','fund','dept',
                     'deptname','bbfe',
                     'allocation_code','pledge_number','donor_id',
                     'pledge_type','pledge_amount','pledge_amount_paid',
                     'date_of_record',
                     'payment_date','payment_amount']
        self.discount_rates={}
        self.query_discount_rates()
        self.pledges={}
        #
        self.query_main_data_and_intake()
        self.process()
        
    def query_discount_rates(self):
        print('querying discount rates')
        lib=library.Library()
        sql=lib.sql_discount_rates()
        for row in lib.download(sql):
            yyyymm,y,m,r,_=row
            k=y,m
            self.discount_rates[k]=r
            
    def query_main_data_and_intake(self):
        lib=library.Library()
        sql=lib.sql_full(pledgemod.EDATE) if ALL_DATA else lib.sql_truncated(pledgemod.EDATE)
        print('querying main data')        
        for row in lib.download(sql):
            adict={k: v for k, v in zip(self.header,row)}
            k=adict['pledge_number']
            #setup
            if not k in self.pledges:
                pledge=pledgemod.PledgeExcel(self.discount_rates)
                self.pledges[k]=pledge
            #intake
            self.pledges[k].intake(adict)
        
    def process(self):
        print('processing pledges')
        for p in self.pledges.values(): p.process()

    def to_excel(self):
        print('writing to excel in %s'%OUTPUT)
        wb=opx.Workbook()
        wb.remove(wb.active)
        self.write_errors(wb)
        self.write_summary(wb)
        self.write_reconciliation(wb)
        self.write_je(wb)
        wb.save(filename=OUTPUT)
        os.startfile(OUTPUT)
    def write_errors(self,wb):
        ws=wb.create_sheet('errors')
        errors=[errors for errors in [pledge.errors_to_excel() for pledge in self.pledges.values()] if errors!=None]
        flattened=list(itertools.chain(*errors))
        flattened.insert(0,'description')
        for i,v in enumerate(flattened):
            ws.cell(row=i+1,column=1).value=v
            if type(v)==type(ZERO) or type(v)==type(0):
                ws.cell(row=i+1,column=1).number_format='0.00'
    def write_summary(self,wb):
        header=['pledge_number','fund','donor_id','pledge_date','pledge_amount','pledge_balance',
        'gross receivable <1y','gross receivable 1y<=x<=5y','gross receivable >5y',
        'haircuts <1y','haircuts 1y<=x<=5y','haircuts >5y',
        'allowances <1y','allowances 1y<=x<=5y','allowances >5y',
                
        'net of haircuts and allowances <1y','net of haircuts and allowances 1y<=x<=5y','net of haircuts and allowances >5y',
        'discounts <1y','discounts 1y<=x<=5y','discounts >5y',
        'net <1y','net 1y<=x<=5y','net >5y',
        ]
        ws=wb.create_sheet('summary')
        summary=[summary for summary in [pledge.summary_to_excel() for pledge in self.pledges.values()]]
        summary.insert(0,header)
        for i,row in enumerate(summary):
            for j,v in enumerate(row):
                ws.cell(row=i+1,column=j+1).value=v
                if type(v)==type(ZERO) or type(v)==type(0):
                        ws.cell(row=i+1,column=j+1).number_format='0.00'
    def write_reconciliation(self,wb):
        header=[['pledge_number','fund','account_number','advance','bbfe','diff']]
        ws=wb.create_sheet('reconciliation')
        recon=[recon for recon in [pledge.reconciliation_to_excel() for pledge in self.pledges.values()]]
        recon.insert(0,header)
        rno=1
        for i,row in enumerate(recon):
            for j,v in enumerate(row):
                for k,v2 in enumerate(v):
                    ws.cell(rno,column=k+1).value=v2
                    if type(v2)==type(ZERO) or type(v2)==type(0):
                        ws.cell(rno,column=k+1).number_format='0.00'
                rno+=1
    def write_je(self,wb):
        header=[['fund','account','post date','journal reference','amount','type','journal','encumbrance','attribute name','attribute value']]
        ws=wb.create_sheet('je')
        recon=[recon for recon in [pledge.je_to_excel() for pledge in self.pledges.values()]]
        recon.insert(0,header)
        rno=1
        for i,row in enumerate(recon):
            for j,v in enumerate(row):
                for k,v2 in enumerate(v):
                    ws.cell(rno,column=k+1).value=v2
                    if type(v2)==type(ZERO) or type(v2)==type(0):
                        ws.cell(rno,column=k+1).number_format='0.00'
                rno+=1
    



if __name__=='__main__':
    app=App()
    app.to_excel()

